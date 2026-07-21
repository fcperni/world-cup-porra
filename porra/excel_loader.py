"""Extracción de datos desde ``docs/ADMIN.xlsx``.

El Excel es la **única fuente de solo lectura** de: selecciones, calendario,
predicciones de los 19 jugadores, reglas de puntuación y cuadro de brackets.
Los resultados introducidos viven aparte (ver :mod:`porra.results_store`).

Mapa de la hoja ADMIN (verificado contra el fichero):

* Filas 6-77   → predicciones de los 72 partidos de grupos (vinculadas a
  WORLDCUP por la fórmula de la columna O ``=WORLDCUP!AC<fila>``).
* Filas 80-127 → posiciones de grupo (12 grupos × 4 posiciones).
* Filas 130-161 → 32 selecciones clasificadas a 1/16.
* Filas 164-179 → 16 partidos de 1/16.
* Filas 182-197 → 16 clasificados a 1/8;  200-207 → 8 partidos de 1/8.
* Filas 210-217 → 8 clasificados a 1/4;   220-223 → 4 partidos de 1/4.
* Filas 226-229 → 4 clasificados a 1/2;   232-233 → 2 partidos de 1/2.
* Filas 236-237 → clasificados al 3-4;    244 → partido 3º-4º.
* Filas 240-241 → finalistas;             247 → final.
* Filas 250-258 → cuadro de honor.

Columnas de jugador: la predicción del jugador *i* está en la columna
``19 + 3*i`` (S, V, Y, …) con el nombre en la fila 5; los puntos cacheados,
en la columna siguiente (que ignoramos: los recalculamos en Python).
"""

from __future__ import annotations

import re
import warnings
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl

from .models import (
    KO_ORDER,
    WC2026_FORMAT,
    KnockoutPrediction,
    Match,
    Phase,
    Player,
    Prediction,
    ScoringRules,
    Team,
    TournamentData,
)
from .venues import venue_for

DEFAULT_XLSX = Path(__file__).resolve().parent.parent / "docs" / "ADMIN.xlsx"

# Columnas (1-based) de la hoja WORLDCUP.
WC_DATE, WC_MATCHDAY, WC_HOME, WC_HG, WC_AG, WC_AWAY, WC_NUM, WC_GROUP = 24, 26, 27, 29, 30, 32, 34, 36

# Columnas (1-based) de la hoja ADMIN.
ADM_RULE_DESC, ADM_RULE_PTS = 3, 4   # C, D
ADM_BONUS, ADM_PHASE, ADM_NAME, ADM_RESULT = 9, 10, 11, 13  # I, J, K, M
ADM_FORMULA_O = 15  # O — fórmula =WORLDCUP!AC<fila>

# Primera columna de predicción de jugador y paso entre jugadores.
PLAYER_COL_START, PLAYER_COL_STEP, N_PLAYERS = 19, 3, 19

# Secciones de filas de la hoja ADMIN.
GROUP_MATCH_ROWS = range(6, 78)        # 72 partidos
GROUP_POS_ROWS = range(80, 128)        # 48 posiciones
QUALIFIED_SECTIONS = {                 # fase de destino -> filas de selecciones clasificadas
    Phase.R32: range(130, 162),
    Phase.R16: range(182, 198),
    Phase.QF: range(210, 218),
    Phase.SF: range(226, 230),
    Phase.THIRD: range(236, 238),
    Phase.FINAL: range(240, 242),
}
KO_MATCH_SECTIONS = {                  # fase -> filas de partidos KO
    Phase.R32: range(164, 180),
    Phase.R16: range(200, 208),
    Phase.QF: range(220, 224),
    Phase.SF: range(232, 234),
    Phase.THIRD: range(244, 245),
    Phase.FINAL: range(247, 248),
}
HONOR_ROWS = {  # fila -> clave del cuadro de honor
    250: "campeon", 251: "subcampeon", 252: "tercero",
    253: "bota_oro", 254: "bota_plata", 255: "bota_bronce",
    256: "balon_oro", 257: "balon_plata", 258: "balon_bronce",
}

# Tokens de fase (sobre la descripción normalizada: mayúsculas, sin espacios).
# El texto del Excel es inconsistente ("OCTAVOS -Diferencia", "3ºy4º PUESTO"),
# por eso normalizamos y comparamos por prefijo.
_PHASE_TOKENS = [
    ("DIECISEISAVOS", Phase.R32),
    ("OCTAVOS", Phase.R16),
    ("CUARTOS", Phase.QF),
    ("SEMIFINALES", Phase.SF),
    ("FASEDEGRUPOS", Phase.GROUPS),
    ("3ºY4º", Phase.THIRD),
    ("FINAL", Phase.FINAL),
]


def _norm(text: str) -> str:
    return "".join(str(text).split()).upper()


# ---------------------------------------------------------------------------
# Parsers de cadenas de predicción
# ---------------------------------------------------------------------------

def parse_score(text: str) -> tuple[Optional[str], Optional[int], Optional[int]]:
    """Parsea ``"<signo>|<local>-<visitante>"`` → (signo, local, visitante).

    Devuelve ``(None, None, None)`` si la cadena no es una predicción válida.
    """
    if not text or "|" not in text:
        return None, None, None
    sign, _, rest = text.partition("|")
    sign = sign.strip()
    m = re.match(r"\s*(\d+)\s*-\s*(\d+)\s*$", rest)
    if not m or sign not in {"1", "X", "2"}:
        return None, None, None
    return sign, int(m.group(1)), int(m.group(2))


def parse_group_prediction(match_number: int, raw) -> Prediction:
    raw = "" if raw is None else str(raw)
    sign, hg, ag = parse_score(raw)
    return Prediction(match_number=match_number, raw=raw, sign=sign, home_goals=hg, away_goals=ag)


def parse_ko_prediction(match_number: int, raw) -> KnockoutPrediction:
    """Parsea ``"<Local>-<Visitante>·<signo>|<local>-<visitante>"``.

    Ningún nombre de selección contiene guion, por lo que la pareja se separa
    de forma inequívoca por el primer ``-``.
    """
    raw = "" if raw is None else str(raw)
    home_team = away_team = sign = hg = ag = None
    if "·" in raw:
        matchup, _, score = raw.partition("·")
        if "-" in matchup:
            home_team, _, away_team = matchup.partition("-")
            home_team, away_team = home_team.strip() or None, away_team.strip() or None
        sign, hg, ag = parse_score(score)
    else:
        sign, hg, ag = parse_score(raw)
    return KnockoutPrediction(
        match_number=match_number, raw=raw, home_team=home_team, away_team=away_team,
        sign=sign, home_goals=hg, away_goals=ag,
    )


def split_placeholder_label(label: str) -> tuple[str, str]:
    """``"1E-3ABCDF"`` → ``("1E", "3ABCDF")``. Divide por el primer guion."""
    home, _, away = label.partition("-")
    return home.strip(), away.strip()


def _as_int(value) -> Optional[int]:
    """Convierte a int si es un número/entero textual; si no, None."""
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return None


# ---------------------------------------------------------------------------
# Carga principal
# ---------------------------------------------------------------------------

def load_tournament(xlsx_path: Path | str = DEFAULT_XLSX) -> TournamentData:
    """Lee ADMIN.xlsx y devuelve un :class:`TournamentData` completo."""
    xlsx_path = Path(xlsx_path)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb_v = openpyxl.load_workbook(xlsx_path, data_only=True)
        wb_f = openpyxl.load_workbook(xlsx_path, data_only=False)

    teams = _load_teams(wb_v["Equipos"])
    matches, wc_row_to_num = _load_matches(wb_v["WORLDCUP"])
    bracket = _load_bracket(wb_v["WORLDCUP"])
    rules = _load_rules(wb_v["ADMIN"])

    # El grupo solo aparece en la 1ª fila de cada bloque de WORLDCUP; para los
    # partidos de grupos lo derivamos del equipo local (ambos comparten grupo).
    team_group = {t.name: t.group for t in teams}
    for m in matches:
        if m.phase is Phase.GROUPS:
            m.group = team_group.get(m.home, m.group)

    # sede (ciudad + estadio) fija por número de partido del calendario FIFA
    for m in matches:
        m.city, m.stadium = venue_for(m.number)

    _attach_admin_metadata(wb_v["ADMIN"], wb_f["ADMIN"], matches, wc_row_to_num)
    players = _load_players(wb_v["ADMIN"], matches)
    thirds_table = _load_thirds_table(wb_v["Combinaciones3"])

    return TournamentData(teams=teams, matches=matches, players=players, rules=rules,
                          bracket=bracket, thirds_table=thirds_table,
                          format=WC2026_FORMAT, competition_id="wc2026")


def load_scoring_rules(xlsx_path: Path | str = DEFAULT_XLSX) -> ScoringRules:
    """Solo la tabla de puntuación de ADMIN.xlsx (fuente única de los valores).

    La reutilizan otras competiciones que aún no tienen su propio Excel (p.ej. el
    stub de la Euro 2028), para partir de una estructura de puntos idéntica a la
    del Mundial.
    """
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(Path(xlsx_path), data_only=True)
    return _load_rules(wb["ADMIN"])


def _load_thirds_table(ws) -> dict[str, dict[str, str]]:
    """Tabla oficial de asignación de mejores terceros (hoja Combinaciones3).

    Columna 14 = combinación de grupos clasificados (``"EFGHIJKL"``).
    Columnas 16-23 = ranuras de primeros (cabecera ``"1A".."1L"``) con el tercero
    asignado (``"3E"``…). Devolvemos ``{combo: {"1A": "E", ...}}``.
    """
    # cabeceras de las 8 ranuras de primero
    slot_headers = {}
    for c in range(16, 24):
        h = ws.cell(1, c).value
        if isinstance(h, str) and h.startswith("1"):
            slot_headers[c] = h.strip()

    table: dict[str, dict[str, str]] = {}
    for r in range(2, ws.max_row + 1):
        combo = ws.cell(r, 14).value
        if not isinstance(combo, str) or not combo.strip():
            continue
        combo = combo.strip()
        mapping = {}
        for c, slot in slot_headers.items():
            val = ws.cell(r, c).value  # p.ej. "3E"
            if isinstance(val, str) and val.startswith("3") and len(val) >= 2:
                mapping[slot] = val[1:].strip()  # letra de grupo
        if mapping:
            table[combo] = mapping
    return table


def _load_teams(ws) -> list[Team]:
    teams = []
    for r in range(2, ws.max_row + 1):
        num = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        if num is None or name in (None, ""):
            continue
        teams.append(Team(num=int(num), name=str(name), group=str(ws.cell(r, 3).value),
                          rank=int(ws.cell(r, 4).value)))
    return teams


def _phase_for_match_number(n: int) -> Phase:
    if n <= 72:
        return Phase.GROUPS
    if n <= 88:
        return Phase.R32
    if n <= 96:
        return Phase.R16
    if n <= 100:
        return Phase.QF
    if n <= 102:
        return Phase.SF
    if n == 103:
        return Phase.THIRD
    return Phase.FINAL


def _load_matches(ws) -> tuple[list[Match], dict[int, int]]:
    """Lee el calendario de WORLDCUP. Devuelve (matches, {fila_wc: nº_partido})."""
    matches: dict[int, Match] = {}
    wc_row_to_num: dict[int, int] = {}
    for r in range(4, ws.max_row + 1):
        num = _as_int(ws.cell(r, WC_NUM).value)
        home = ws.cell(r, WC_HOME).value
        away = ws.cell(r, WC_AWAY).value
        if num is None or home in (None, "") or away in (None, ""):
            continue
        wc_row_to_num[r] = num
        group_raw = ws.cell(r, WC_GROUP).value
        group = group_raw[-1] if isinstance(group_raw, str) and group_raw.startswith("Grupo") else None
        matchday = ws.cell(r, WC_MATCHDAY).value
        matchday = matchday if isinstance(matchday, str) and matchday.startswith("J") else None
        date = ws.cell(r, WC_DATE).value
        matches[num] = Match(
            number=num,
            phase=_phase_for_match_number(num),
            matchday=matchday,
            group=group,
            home=str(home),
            away=str(away),
            date=date if isinstance(date, datetime) else None,
        )
    return [matches[n] for n in sorted(matches)], wc_row_to_num


def _load_bracket(ws) -> dict[int, tuple[str, str]]:
    """nº de partido KO -> (placeholder_local, placeholder_visitante)."""
    bracket = {}
    for r in range(4, ws.max_row + 1):
        num = _as_int(ws.cell(r, WC_NUM).value)
        if num is None or num <= 72:
            continue
        home, away = ws.cell(r, WC_HOME).value, ws.cell(r, WC_AWAY).value
        if home in (None, "") or away in (None, ""):
            continue
        bracket[num] = (str(home), str(away))
    return bracket


def _load_rules(ws) -> ScoringRules:
    points: dict = {}
    rows = []  # (desc_original, desc_normalizada, puntos)
    for r in range(8, 48):
        desc = ws.cell(r, ADM_RULE_DESC).value
        pts = ws.cell(r, ADM_RULE_PTS).value
        if desc and pts is not None:
            rows.append((str(desc), _norm(desc), float(pts)))

    # categorías por partido: signo / diferencia / exacto, detectando fase por prefijo
    for desc, norm, pts in rows:
        if norm.startswith("EQUIPOCLASIFICADO") or norm.startswith("POSICIÓNEXACTA"):
            continue
        phase = next((ph for tok, ph in _PHASE_TOKENS if norm.startswith(tok)), None)
        if phase is None:
            continue
        if "SIGNO" in norm:
            points[(phase, "signo")] = pts
        elif "DIFERENCIA" in norm or "DISTANCIA" in norm:
            points[(phase, "diferencia")] = pts
        elif "RESULTADOEXACTO" in norm:
            points[(phase, "exacto")] = pts

    # "Equipo clasificado para X" -> puntos por acertar clasificado a esa fase
    qualif_map = {
        "DIECISEISAVOS": Phase.R32, "OCTAVOS": Phase.R16, "CUARTOS": Phase.QF,
        "SEMIFINALES": Phase.SF, "3ºY4º": Phase.THIRD, "FINAL": Phase.FINAL,
    }
    for desc, norm, pts in rows:
        if norm.startswith("EQUIPOCLASIFICADO"):
            for token, phase in qualif_map.items():
                if token in norm:
                    points[(phase, "clasificado")] = pts
                    break

    # posiciones exactas de grupo (5 c/u)
    for desc, norm, pts in rows:
        if "POSICIÓNEXACTA" in norm:
            m = re.search(r"\((\d)", desc)
            if m:
                points[("group_position", int(m.group(1)))] = pts

    # cuadro de honor (filas 39-47)
    honor_labels = {
        "campeon": "CAMPEÓN", "subcampeon": "SUBCAMPEÓN", "tercero": "3ºPUESTO",
        "bota_oro": "BOTADEORO", "bota_plata": "BOTADEPLATA", "bota_bronce": "BOTADEBRONCE",
        "balon_oro": "BALÓNDEORO", "balon_plata": "BALÓNDEPLATA", "balon_bronce": "BALÓNDEBRONCE",
    }
    for desc, norm, pts in rows:
        for hk, label in honor_labels.items():
            if norm.startswith(label):
                points[("honor", hk)] = pts
                break

    # factor de ajuste por diferencia (D50)
    diff_adjust = ws.cell(50, ADM_RULE_PTS).value
    diff_adjust = float(diff_adjust) if isinstance(diff_adjust, (int, float)) else 0.1
    return ScoringRules(points=points, diff_adjust=diff_adjust)


def _attach_admin_metadata(ws_v, ws_f, matches: list[Match], wc_row_to_num: dict[int, int]) -> None:
    """Asocia a cada Match su fila de ADMIN y su bonus."""
    by_num = {m.number: m for m in matches}

    # Grupos: la fórmula O de cada fila apunta a WORLDCUP!AC<fila_wc>.
    pat = re.compile(r"WORLDCUP!AC(\d+)")
    for r in GROUP_MATCH_ROWS:
        formula = ws_f.cell(r, ADM_FORMULA_O).value
        bonus = ws_v.cell(r, ADM_BONUS).value
        if not isinstance(formula, str):
            continue
        m = pat.search(formula)
        if not m:
            continue
        wc_row = int(m.group(1))
        num = wc_row_to_num.get(wc_row)
        if num and num in by_num:
            by_num[num].admin_row = r
            by_num[num].bonus = int(bonus) if isinstance(bonus, (int, float)) else 1

    # Eliminatorias: la etiqueta K ("1E-3ABCDF") identifica el cruce por placeholders.
    placeholder_to_num = {}
    for m in matches:
        if m.phase.is_knockout:
            placeholder_to_num[(m.home, m.away)] = m.number
    for phase, rows in KO_MATCH_SECTIONS.items():
        for r in rows:
            label = ws_v.cell(r, ADM_NAME).value
            bonus = ws_v.cell(r, ADM_BONUS).value
            if not isinstance(label, str):
                continue
            home, away = split_placeholder_label(label)
            num = placeholder_to_num.get((home, away))
            if num and num in by_num:
                by_num[num].admin_row = r
                by_num[num].bonus = int(bonus) if isinstance(bonus, (int, float)) else 1


def _load_players(ws, matches: list[Match]) -> list[Player]:
    players: list[Player] = []
    # mapas fila_admin -> nº de partido
    group_row_to_num = {m.admin_row: m.number for m in matches
                        if m.phase is Phase.GROUPS and m.admin_row}
    ko_row_to_num = {m.admin_row: m.number for m in matches
                     if m.phase.is_knockout and m.admin_row}

    for i in range(N_PLAYERS):
        col = PLAYER_COL_START + i * PLAYER_COL_STEP
        name = ws.cell(5, col).value
        if name in (None, ""):
            continue
        p = Player(name=str(name).strip(), column=col)

        # grupos
        for r, num in group_row_to_num.items():
            p.group_matches[num] = parse_group_prediction(num, ws.cell(r, col).value)

        # posiciones de grupo (filas 80-127)
        for r in GROUP_POS_ROWS:
            idx = r - GROUP_POS_ROWS.start
            group = chr(ord("A") + idx // 4)
            position = idx % 4 + 1
            val = ws.cell(r, col).value
            p.group_positions[(group, position)] = str(val).strip() if val not in (None, "") else None

        # clasificados por fase
        for phase, rows in QUALIFIED_SECTIONS.items():
            teams = []
            for r in rows:
                val = ws.cell(r, col).value
                if val not in (None, ""):
                    teams.append(str(val).strip())
            p.qualified[phase] = teams

        # partidos KO
        for r, num in ko_row_to_num.items():
            p.ko_matches[num] = parse_ko_prediction(num, ws.cell(r, col).value)

        # cuadro de honor
        for r, key in HONOR_ROWS.items():
            val = ws.cell(r, col).value
            p.honor[key] = str(val).strip() if val not in (None, "") else None

        players.append(p)
    return players


if __name__ == "__main__":  # verificación rápida de extracción
    data = load_tournament()
    print(f"Equipos: {len(data.teams)} (esperado 48)")
    groups = sorted({t.group for t in data.teams})
    print(f"Grupos: {groups}")
    print(f"Partidos: {len(data.matches)} (esperado 104)")
    by_phase = {}
    for m in data.matches:
        by_phase[m.phase.value] = by_phase.get(m.phase.value, 0) + 1
    print(f"Partidos por fase: {by_phase}")
    linked = sum(1 for m in data.matches if m.admin_row)
    print(f"Partidos con fila ADMIN vinculada: {linked} (esperado 104)")
    print(f"Jugadores: {len(data.players)} -> {[p.name for p in data.players]}")
    print(f"Reglas extraídas: {len(data.rules.points)} entradas; ajuste dif={data.rules.diff_adjust}")
    print(f"Bracket KO entradas: {len(data.bracket)} (esperado 32)")

    p = data.players[7]  # PACO
    print(f"\nEjemplo {p.name}:")
    m1 = data.matches[0]
    print(f"  Partido {m1.number} {m1.home}-{m1.away} (bonus {m1.bonus}) -> pred {p.group_matches[m1.number].raw!r}")
    print(f"  Posiciones grupo A: " + ", ".join(
        f"{pos}º={p.group_positions[('A', pos)]}" for pos in range(1, 5)))
    print(f"  Clasificados 1/16: {len(p.qualified[Phase.R32])} equipos; finalistas: {p.qualified[Phase.FINAL]}")
    print(f"  Honor: {p.honor}")
    ko = p.ko_matches[104]  # la final es el partido nº 104
    print(f"  Final (p104): {ko.raw!r} -> {ko.home_team} vs {ko.away_team}, sign={ko.sign} {ko.home_goals}-{ko.away_goals}")
